import tkinter

import pyOTDR
import os
import re
from openpyxl import Workbook



pliksor = re.compile(r'.sor')
fala = re.compile(r'1310')
drop = re.compile(r'drop')
wb = Workbook()
ws = wb.active
ws['A1'] = "Złącze numer"
ws['H1'] = "Złącze numer"
ws['B1'] = "[dB]"
ws['I1'] = "[dB]"
ws['C1'] = "[km]"
ws['J1'] = "[km]"
ws['F1'] = "Złącze A"
ws['M1'] = "Złącze A"

baz_fold = input("podaj całą ściężkę\n")
nazwapliku = input("Podaj nazwę pliku końcowego\n")
licz = 0
licz2 = 2

sub_folders = []
for dir, sub_dirs, files in os.walk(baz_fold):
    sub_folders.extend(sub_dirs)
    sub_folders1 = str(f"{dir}")
    # print(f'foldery {sub_folders1}')

    for plik in os.listdir(sub_folders1):
        if pliksor.search(plik):
            # print(f'nazwa pliku {plik}')
            if fala.search(plik):
                new = pyOTDR.sorparse(str(sub_folders1) + '\\' + str(plik))
                new2 = new[1]
                nazwa2 =plik.split(sep='_')
                print(new2["GenParams"]["location A"])
                print(nazwa2)
                nazwa3 = nazwa2[1]
                if nazwa3 == "DROP":
                    nazwa = new2["GenParams"]["location A"] + f' P'+nazwa2[2]
                else:
                    nazwa = new2["GenParams"]["location A"] + f' SP'+ nazwa2[2]
                # print(nazwa2)
                db = new2["KeyEvents"]['Summary']["total loss"]
                lenght = new2["KeyEvents"]['Summary']["ORL finish"]
                ws.cell(row=licz2, column=1).value = nazwa
                wavelenght = new2["GenParams"]["wavelength"]
                ws.cell(row=licz2, column=2).value = db
                ws.cell(row=licz2, column=3).value = lenght
                if new2["KeyEvents"]['event 1']["refl loss"] != "0.000":
                    ws.cell(row=licz2, column=6).value = new2["KeyEvents"]['event 1']["refl loss"]
                licz2 -= 1
            else:
                new = pyOTDR.sorparse(str(sub_folders1) + '\\' + str(plik))
                new2 = new[1]
                nazwa2 = plik.split(sep='_')
                # print(nazwa2)
                nazwa3 = nazwa2[1]
                if nazwa3 == "DROP":
                    nazwa = new2["GenParams"]["location A"] + f' P' + nazwa2[2]
                else:
                    nazwa = new2["GenParams"]["location A"] + f' SP' + nazwa2[2]
                # print(plik)
                db = new2["KeyEvents"]['Summary']["total loss"]
                lenght = new2["KeyEvents"]['Summary']["ORL finish"]
                ws.cell(row=licz2, column=8).value = nazwa
                wavelenght = new2["GenParams"]["wavelength"]
                ws.cell(row=licz2, column=9).value = db
                ws.cell(row=licz2, column=10).value = lenght
                if new2["KeyEvents"]['event 1']["refl loss"] != "0.000":
                    ws.cell(row=licz2, column=13).value = new2["KeyEvents"]['event 1']["refl loss"]
            licz += 1
            licz2 += 1
wb.save(f"{baz_fold}\{nazwapliku}.xlsx")
print(licz)
