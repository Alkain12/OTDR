# from glob import glob
import tkinter as tk
import tkinter.ttk as ttk
from tkinter.ttk import Progressbar
from tkinter import filedialog
import pyOTDR
import os
import re
from openpyxl import Workbook
import tkinter
# import sys

root = tkinter.Tk()
root.title("Pliki Sor do XLSX")
root.geometry('650x380+100+100')

pliksor = re.compile(r'.sor')
fala = re.compile(r'1310')
drop = re.compile(r'drop')
wb = Workbook()
ws = wb.active
ws['A1'] = "Złącze numer"
ws['H1'] = "Złącze numer"
ws['B1'] = "[dB]"
ws['I1'] = "[dB]"
ws['C1'] = "[km]"
ws['J1'] = "[km]"
ws['F1'] = "Złącze A"
ws['M1'] = "Złącze A"

# baz_fold = input("podaj całą ściężkę\n")
baz_fold = str()
# nazwapliku = input("Podaj nazwę pliku końcowego\n")
nazwapliku = str()
licz = 0
liczpliki = 0
var = tk.IntVar()


def miejsce_button():
    def f():
        miejsce_butt: str = tkinter.filedialog.askdirectory()
        miejsce.insert(-1, miejsce_butt)
    return f()


def pobranie():
    def f():
        miejsce2 = miejsce.get()
        plik2 = plik.get()
        nazwa_czlon_2_licz = nazwa_czlon_2.get()
        nazwa2_czlon_2_licz = nazwa2_czlon_2.get()
        opis_czlon_2_licz = opis_czlon_2.get()
        opis2_czlon_2_licz = opis2_czlon_2.get()
        opis2_czlon_2c_licz = opis2_czlon_2c.get()
        opis2_refelktancja_licz = float(opis2_refelktancja.get())
        # print(miejsce2)
        # print(plik2)
        liczenie(miejsce2, plik2, nazwa_czlon_2_licz,nazwa2_czlon_2_licz, opis_czlon_2_licz, opis2_czlon_2_licz, opis2_czlon_2c_licz, opis2_refelktancja_licz)

    return f()


def liczenie(baz_fold, nazwapliku, second_czlon_drop, second_czlon_olt, opis_z_d, opis_z_olt, opis_z_pozostale, opis_z_reflektancja):
    licz_all = 0
    licz2 = 2

    sub_folders = []
    for dir, sub_dirs, files in os.walk(baz_fold):
        sub_folders.extend(sub_dirs)
        sub_folders1 = str(f"{dir}")
        # print(f'foldery {sub_folders1}')

        for plik in os.listdir(sub_folders1):
            if pliksor.search(plik):
                licz_all += 1
                # print(licz_all)
    licz_all_10_pr : int = licz_all // 10
    # print(licz_all_10_pr)
    licz = 0
    # licztdqm = 0
    sub_folders = []
    for dir, sub_dirs, files in os.walk(baz_fold):
        sub_folders.extend(sub_dirs)
        sub_folders1 = str(f"{dir}")
        # print(f'foldery {sub_folders1}')

        for plik in os.listdir(sub_folders1):
            if pliksor.search(plik):

                # print(f'nazwa pliku {plik}')
                if fala.search(plik):
                    open_sor = pyOTDR.sorparse(str(sub_folders1) + '\\' + str(plik))
                    new2 = open_sor[1]
                    nazwa2 = plik.split(sep='_')
                    reflektancja = float(new2["KeyEvents"]['event 1']["refl loss"])
                    # print(new2["GenParams"])
                    nazwa3 = nazwa2[1]
                    if nazwa3 == second_czlon_drop:
                        nazwa = new2["GenParams"]["location A"] + " " + opis_z_d + nazwa2[2]
                    if nazwa3 == second_czlon_olt:
                        nazwa = new2["GenParams"]["location A"] + " " + opis_z_olt + nazwa2[2]
                    if nazwa3 != second_czlon_drop and nazwa3 != second_czlon_olt:
                        nazwa = new2["GenParams"]["location B"] + " " + opis_z_pozostale + nazwa2[2]
                    # print(nazwa2)
                    db = new2["KeyEvents"]['Summary']["total loss"]
                    lenght = new2["KeyEvents"]['Summary']["ORL finish"]
                    ws.cell(row=licz2, column=1).value = nazwa
                    # wavelenght = new2["GenParams"]["wavelength"]
                    ws.cell(row=licz2, column=2).value = db
                    ws.cell(row=licz2, column=3).value = lenght
                    if new2["KeyEvents"]['event 1']["refl loss"] != "0.000":
                        if reflektancja < opis_z_reflektancja:
                            ws.cell(row=licz2, column=6).value = ""
                        else:
                            #print(reflektancja," =>" ,opis_z_reflektancja)
                            ws.cell(row=licz2, column=6).value = new2["KeyEvents"]['event 1']["refl loss"]
                    licz2 -= 1
                else:
                    open_sor = pyOTDR.sorparse(str(sub_folders1) + '\\' + str(plik))
                    new2 = open_sor[1]
                    nazwa2 = plik.split(sep='_')
                    reflektancja = float(new2["KeyEvents"]['event 1']["refl loss"])
                    # print(nazwa2)
                    nazwa3 = nazwa2[1]
                    if nazwa3 == second_czlon_drop:
                        nazwa = new2["GenParams"]["location A"] + " " + opis_z_d + nazwa2[2]
                    if nazwa3 == second_czlon_olt:
                        nazwa = new2["GenParams"]["location A"] + " " + opis_z_olt + nazwa2[2]
                    if nazwa3 != second_czlon_drop and nazwa3 != second_czlon_olt:
                        nazwa = new2["GenParams"]["location B"] + f' P' + nazwa2[2]
                    # print(plik)
                    db = new2["KeyEvents"]['Summary']["total loss"]
                    lenght = new2["KeyEvents"]['Summary']["ORL finish"]
                    ws.cell(row=licz2, column=8).value = nazwa
                    # wavelenght = new2["GenParams"]["wavelength"]
                    ws.cell(row=licz2, column=9).value = db
                    ws.cell(row=licz2, column=10).value = lenght
                    if new2["KeyEvents"]['event 1']["refl loss"] != "0.000":
                        if reflektancja < opis_z_reflektancja:
                            ws.cell(row=licz2, column=13).value = ""
                        else:
                            # print(reflektancja," =>" ,opis_z_reflektancja)
                            ws.cell(row=licz2, column=13).value = new2["KeyEvents"]['event 1']["refl loss"]
                licz += 1
                licz2 += 1


    wb.save(f"{baz_fold}\{nazwapliku}.xlsx")
    # print(licz)
    new_wind = tkinter.Toplevel(root)
    new_wind.geometry('200x200')
    new_wind.title("Koniec pracy")
    tkinter.Label(new_wind, text=f"Skończone przetworzone {licz} plików").pack()
    # sys.exit()


tkinter.Label(root, text='Gdzie są pliki?').grid(row=0, column=0, padx=2, pady=2)
b1 = tkinter.Button(root, text='Wskaż miejsce ', command=miejsce_button).grid(row=0, column=2, pady=2, padx=2)
miejsce = tkinter.Entry(root)
# b1 = tkinter.Button(root, text='Wskaż miejsce ', command=miejsce_button).grid(row=0, column=2,pady=2,padx=2)
miejsce.grid(row=0, column=1)

tkinter.Label(root, text='2 człon nazwy pliku').grid(row=1, column=0, padx=2, pady=2)
nazwa_czlon_2 = tkinter.Entry(root)
nazwa_czlon_2.grid(row=1, column=1)
nazwa_czlon_2.insert(-1, "DROP")

tkinter.Label(root, text='opis excel').grid(row=1, column=2, padx=2, pady=2)
opis_czlon_2 = tkinter.Entry(root)
opis_czlon_2.grid(row=1, column=3)
opis_czlon_2.insert(-1, "D")

tkinter.Label(root, text='2 człon nazwy pliku').grid(row=2, column=0, padx=2, pady=2)
nazwa2_czlon_2 = tkinter.Entry(root)
nazwa2_czlon_2.grid(row=2, column=1)
nazwa2_czlon_2.insert(-1, "OLT")

tkinter.Label(root, text='opis excel').grid(row=2, column=2, padx=2, pady=2)
opis2_czlon_2 = tkinter.Entry(root)
opis2_czlon_2.grid(row=2, column=3)
opis2_czlon_2.insert(-1, "SP")

tkinter.Label(root, text='Pozostałe pliki').grid(row=3, column=0, padx=2, pady=2)
nazwa2_czlon_2c = tkinter.Entry(root)
#nazwa2_czlon_2.grid(row=3, column=1)
#nazwa2_czlon_2.insert(-1, "OLT")

tkinter.Label(root, text='opis excel').grid(row=3, column=2, padx=2, pady=2)
opis2_czlon_2c = tkinter.Entry(root)
opis2_czlon_2c.grid(row=3, column=3)
opis2_czlon_2c.insert(-1, "P")

tkinter.Label(root, text='reflektancja').grid(row=4, column=2, padx=2, pady=2)
opis2_refelktancja = tkinter.Entry(root)
opis2_refelktancja.grid(row=4, column=3)
opis2_refelktancja.insert(-1, "-72")



tkinter.Label(root, text='Nazwa pliku xlsx').grid(row=6, column=0, padx=2, pady=2)
plik = tkinter.Entry(root)
plik.grid(row=6, column=1)

b2 = tkinter.Button(root, text='Twórz XLSX, plik ', command=pobranie).grid(row=6)


style = ttk.Style()
style.theme_use('default')
style.configure("black.Horizontal.TProgressbar", background='brown')

pasek = Progressbar(root, length=500,  style='black.Horizontal.TProgressbar')
pasek.grid(row=7, column=0, columnspan=4, padx=2, pady=2)
# pasek['value'] = 0

root.mainloop()
